from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import IntegrityError
from django.http import HttpResponseForbidden
from django.urls import reverse
from django import forms
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
# Models
from .models import Class, SectionGroup, Section, Students, User, ExcelRowData, Booking, Passenger
# Forms
import pandas as pd
import openpyxl
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_protect
from .forms import CustomUserCreationForm
from flightapp.models import Route, Schedule  # adjust app name if different
from flightapp.models import Route   # ✅ import Route from flightapp
from .models import MultiCityLeg
from django.utils.dateparse import parse_date, parse_datetime
from datetime import date
from django.utils.timezone import make_aware
User = get_user_model()
from django.utils import timezone


@csrf_protect   # ✅ ensure CSRF token is required & generated
def registerPage(request):
    success = False
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            success = True
            return redirect("login")  # ✅ redirect after success
    else:
        form = CustomUserCreationForm()

    return render(
        request,
        "tutor/register.html",
        {
            "form": form,
            "errors": form.errors if form.errors else None,
            "success": success,
        },
    )


# Login view
def loginPage(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not User.objects.filter(username=username).exists():
            return render(request, "tutor/login.html", {"error": "Username does not exist", "form": AuthenticationForm()})

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            return render(request, "tutor/login.html", {"error": "Invalid password", "form": AuthenticationForm()})
    else:
        form = AuthenticationForm()
    return render(request, "tutor/login.html", {"form": form})

# Logout view
def logoutPage(request):
    logout(request)
    return redirect("login")

# Create your views here.
@login_required
def dashboard(request):
    return render(request, 'tutor/Dashboard.html')

# Class List

@login_required
def Cclass(request):
    query = request.GET.get('q', '')
    classes_list = Class.objects.filter(user=request.user).order_by('id')  # ✅ only user’s classes

    if query:
        classes_list = classes_list.filter(
            Q(subject_code__icontains=query) |
            Q(subject_title__icontains=query) |
            Q(instructor__icontains=query) |
            Q(schedule__icontains=query) |
            Q(semester__icontains=query) |
            Q(academic_year__icontains=query)
        )

    paginator = Paginator(classes_list, 15)
    page_number = request.GET.get('page')
    classes = paginator.get_page(page_number)

    return render(request, 'tutor/Cclass.html', {
        'classes': classes,
        'query': query,
    })



# Create Class
@login_required
def CreateClass(request):
    if request.method == "POST":
        subject_code = request.POST.get("subject_code")
        subject_title = request.POST.get("subject_title")
        schedule = request.POST.get("schedule")
        semester = request.POST.get("semester")
        academic_year = request.POST.get("academic_year")

        Class.objects.create(
            user=request.user,  # ✅ link to logged in user
            subject_code=subject_code,
            subject_title=subject_title,
            instructor=request.user.username,  # ✅ auto use logged in instructor
            schedule=schedule,
            semester=semester,
            academic_year=academic_year
        )

        messages.success(request, "Successfully Created!")
        return redirect("CreateClass")

    return render(request, "Create/CreateClass.html")

#Edit Class


@login_required
def EditClass(request, id):
    cls = get_object_or_404(Class, id=id, user=request.user)


    if request.method == "POST":
        cls.subject_code = request.POST.get("subject_code")
        cls.subject_title = request.POST.get("subject_title")
        cls.instructor = request.POST.get("instructor") or request.user.get_full_name() or request.user.username
        cls.schedule = request.POST.get("schedule")
        cls.semester = request.POST.get("semester")
        cls.academic_year = request.POST.get("academic_year")
        cls.save()
        return redirect("Cclass")

    return render(request, "Cedit/editclass.html", {"cls": cls})

#Delete Class


@login_required
def DeleteClass(request, id):
    cls = get_object_or_404(Class, id=id, user=request.user)

    cls.delete()
    return redirect('Cclass')







@login_required
def Section_List(request):
    groups = SectionGroup.objects.filter(user=request.user).prefetch_related("sections")

    colors = ["#1976d2", "#f57c00", "#388e3c", "#c2185b", "#6a1b9a", "#00838f", "#d81b60"]

    # Fetch subject codes + titles (unique)
    class_map = {
        c["subject_code"]: c["subject_title"]
        for c in Class.objects.filter(user=request.user)
        .values("subject_code", "subject_title")
        .distinct()
    }

    for idx, group in enumerate(groups):
        group.card_color = colors[idx % len(colors)]
        group.subject_title = class_map.get(group.subject_code, "No Title")

        # Always ensure instructor is set
        if not getattr(group, "instructor", None):
            group.instructor = request.user.username

    return render(request, "tutor/Section_List.html", {"groups": groups})

@login_required
def Create_Section(request):
    if request.method == "POST":
        subject_code = request.POST.get("subject_code")
        num_sections = int(request.POST.get("num_sections"))

        subject_title = Class.objects.filter(user=request.user, subject_code=subject_code).values_list("subject_title", flat=True).first()

        section_group = SectionGroup.objects.create(
            user=request.user,  # ✅
            subject_code=subject_code,
            subject_title=subject_title,
            instructor=request.user.username,  # ✅ auto
            num_sections=num_sections
        )

        for i in range(1, num_sections + 1):
            section_name = request.POST.get(f"section_{i}")
            section_schedule = request.POST.get(f"schedule_{i}")
            if section_name and section_schedule:
                Section.objects.create(group=section_group, name=section_name, schedule=section_schedule)

        return redirect("Section_List")

    subject_codes = Class.objects.filter(user=request.user).values_list("subject_code", flat=True).distinct()
    classes = Class.objects.filter(user=request.user)

    context = {
        "subject_codes": subject_codes,
        "classes": classes,
        "section_range": range(1, 11),
    }
    return render(request, "Create/Create_Section.html", context)





# Update Section View


@login_required
def Update_section(request, pk):
    group = get_object_or_404(SectionGroup, pk=pk)
    sections = group.sections.all()

    # Fetch both subject_code and subject_title
    classes = list(
        Class.objects.filter(user=request.user)
        .values("subject_code", "subject_title")
        .distinct()
    )
    subject_codes = [c["subject_code"] for c in classes]

    if request.method == "POST":
        subject_code = request.POST.get("subject_code")
        group.subject_code = subject_code

        # auto fetch subject title
        subject_title = next((c["subject_title"] for c in classes if c["subject_code"] == subject_code), "")
        group.subject_title = subject_title

        group.instructor = request.user.username
        group.save()

        for section in sections:
            name_key = f"section_name_{section.pk}"
            schedule_key = f"section_schedule_{section.pk}"
            if name_key in request.POST and schedule_key in request.POST:
                section.name = request.POST.get(name_key)
                section.schedule = request.POST.get(schedule_key)
                section.save()

        messages.success(request, "Section group updated successfully.")
        return redirect("Section_List")

    subject_title = next((c["subject_title"] for c in classes if c["subject_code"] == group.subject_code), "")

    return render(request, "Cedit/editsection.html", {
        "group": group,
        "sections": sections,
        "subject_codes": subject_codes,
        "subject_title": subject_title,
        "username": request.user.username,
        "classes": classes,   # ✅ now contains both code + title
    })

# ✅ Delete SectionGroup

@login_required
def Delete_section(request, id):
    group = get_object_or_404(SectionGroup, id=id, user=request.user)

    group.delete()
    return redirect('Section_List')







@login_required
def create_students(request, pk):
    section = get_object_or_404(Section, pk=pk)

    # Get headers dynamically from first uploaded Excel row
    row_objects = ExcelRowData.objects.filter(section=section)
    headers = []
    if row_objects.exists():
        headers = list(row_objects.first().data.keys())

    if request.method == "POST":
        count = int(request.POST.get("count", 1))
        for i in range(1, count + 1):
            # Build dynamic dict from headers
            student_data = {}
            missing_required = False

            for header in headers:
                value = request.POST.get(f"{header}_{i}")
                if not value:
                    missing_required = True
                student_data[header] = value

            if not missing_required:
                ExcelRowData.objects.create(
                    section=section,
                    data=student_data
                )

        return redirect("Section_Detail", pk=pk)

    return render(
        request,
        "Create/Create_students.html",
        {
            "section": section,
            "count_range": range(1, 101),
            "headers": headers,   # Pass headers to HTML
        },
    )


@login_required
def import_students_excel(request, pk):
    section = get_object_or_404(Section, pk=pk)

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]  # first row headers

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            ExcelRowData.objects.create(
                section=section,
                data=row_dict
            )

        return redirect("Section_Detail", pk=section.pk)

    return redirect("tutor/Section_Detail", pk=section.pk)


@login_required
def section_detail(request, pk):
    section = get_object_or_404(Section, pk=pk)
    row_objects = ExcelRowData.objects.filter(section=section)
    
    # Search Excel rows
    query = request.GET.get("q", "").strip()
    if query:
        # SQLite-safe search: filter manually
        row_objects = [row for row in row_objects if query.lower() in str(row.data).lower()]
    
    headers = []
    rows = []
    if row_objects:
        first_row = row_objects[0].data or {}
        headers = list(first_row.keys())
        for obj in row_objects:
            rows.append({"id": obj.id, "values": list(obj.data.values())})
    
    # Fetch Bookings with related data - properly handle route relationships
    bookings = Booking.objects.filter(section=section).prefetch_related(
        'passengers',  # From Passenger model: related_name="passengers"
        'legs'         # From MultiCityLeg model: related_name="legs"
    ).order_by('-created_at')
    
    # Process route information for each booking
    for booking in bookings:
        if booking.route:
            try:
                # Check if route is a numeric ID (database ID) or actual route code
                if booking.route.isdigit():
                    # It's a route ID, fetch the actual route object
                    route_obj = Route.objects.select_related('origin_airport', 'destination_airport').get(id=int(booking.route))
                    # Display as ORIGIN-DESTINATION format
                    booking.route_display = f"{route_obj.origin_airport.code}-{route_obj.destination_airport.code}"
                else:
                    # It's already a route code string, display as is
                    booking.route_display = booking.route
            except (Route.DoesNotExist, ValueError):
                # If route ID doesn't exist or invalid, show the raw value
                booking.route_display = booking.route
        else:
            booking.route_display = "Route not specified"
    
    return render(request, "tutor/Section_Detail.html", {
        "section": section,
        "headers": headers,
        "rows": rows,
        "query": query,
        "bookings": bookings,
        "now": timezone.now(),  # Add current time for duration comparison
    })



@login_required
def delete_excel_row(request, pk):
    row = get_object_or_404(ExcelRowData, pk=pk)
    section_id = row.section.id
    row.delete()
    messages.success(request, "Row deleted successfully!")
    # redirect to the Section_Detail view by name, not to the template file
    return redirect("Section_Detail", pk=section_id)



@login_required
def edit_student(request, pk):
    student = get_object_or_404(ExcelRowData, pk=pk)
    section = student.section

    # Get headers from first Excel row for consistency
    row_objects = ExcelRowData.objects.filter(section=section)
    headers = []
    if row_objects.exists():
        headers = list(row_objects.first().data.keys())

    if request.method == "POST":
        updated_data = {}
        for header in headers:
            updated_data[header] = request.POST.get(header, "")

        student.data = updated_data
        student.save()
        messages.success(request, "Student updated successfully!")
        return redirect("Section_Detail", pk=section.pk)

    return render(request, "Cedit/Edit_student.html", {
        "section": section,
        "student": student,
        "headers": headers,
    })



def book_flight(request, section_id):
    section = get_object_or_404(Section, id=section_id)

    # ✅ fetch all routes
    routes = Route.objects.select_related("origin_airport", "destination_airport").all()

    # ✅ fetch all schedules with related flight + routes
    schedules = Schedule.objects.select_related("flight__route__origin_airport", "flight__route__destination_airport").all()

    if request.method == "POST":
        trip_type = request.POST.get("trip_type", "one_way")
        route_id = request.POST.get("route")
        departure_date = request.POST.get("departure_date")
        return_date = request.POST.get("return_date") or None
        duration = request.POST.get("duration") or None
        travel_class = request.POST.get("travel_class", "economy")
        adults = int(request.POST.get("adults", 1))
        children = int(request.POST.get("children", 0))
        infants_on_lap = int(request.POST.get("infants_on_lap", 0))
        seat_preference = request.POST.get("seat_preference", "")

        if not departure_date:
            return render(request, "Create/Create_Instruction.html", {
                "section": section,
                "routes": routes,
                "schedules": schedules,
                "error": "Departure date is required."
            })

        # ✅ Handle multi-city booking
        multi_city_data = []
        if trip_type == "multi_city":
            i = 1
            while f"from_{i}" in request.POST:
                origin = request.POST.get(f"from_{i}")
                dest = request.POST.get(f"to_{i}")
                date = request.POST.get(f"date_{i}")
                leg_duration = request.POST.get(f"duration_{i}") or None
                if origin and dest and date:
                    multi_city_data.append({
                        "origin": origin,
                        "destination": dest,
                        "date": parse_date(date),
                        "duration": make_aware(parse_datetime(leg_duration)) if leg_duration else None,  # Fix timezone
                    })
                i += 1

        # ✅ Store the original route format for display
        original_route = route_id
        
        # ✅ Handle round trip route parsing - but keep original format for storage
        if trip_type == "round_trip" and route_id and "|" in route_id:
            # Keep the full route for display (e.g., "BXU-CEB|CEB-BXU")
            # You can split for processing if needed, but store the complete route
            outbound_route_id, return_route_id = route_id.split("|")
            # For round trip, you might want to store the complete route or just outbound
            # Keeping original format for better display
        
        # ✅ Create the main booking
        booking = Booking.objects.create(
            section=section,
            trip_type=trip_type,
            route=original_route,  # Store the original route format (like BXU-CEB or BXU-CEB|CEB-BXU)
            departure_date=parse_date(departure_date),
            return_date=parse_date(return_date) if return_date else None,
            duration=make_aware(parse_datetime(duration)) if duration else None,  # Fix timezone warning
            adults=adults,
            children=children,
            infants_on_lap=infants_on_lap,
            travel_class=travel_class,
            seat_preference=seat_preference,
        )

        # ✅ Handle multi-city legs
        if multi_city_data:
            for leg in multi_city_data:
                MultiCityLeg.objects.create(
                    booking=booking,
                    origin=leg["origin"],
                    destination=leg["destination"],
                    departure_date=leg["date"],
                    duration=leg["duration"],
                )

        # ✅ Handle passengers
        passenger_count = int(request.POST.get("passenger_count", 1))
        for p in range(1, passenger_count + 1):
            first_name = request.POST.get(f"first_name_{p}", "").strip()
            last_name = request.POST.get(f"last_name_{p}", "").strip()
            middle_initial = request.POST.get(f"mi_{p}", "").strip() or None

            if not first_name or not last_name:
                continue

            dob_year = request.POST.get(f'year_{p}')
            dob_month = request.POST.get(f'month_{p}')
            dob_day = request.POST.get(f'day_{p}')
            dob = None
            if dob_year and dob_month and dob_day:
                try:
                    dob = parse_date(f"{dob_year}-{dob_month.zfill(2)}-{dob_day.zfill(2)}")
                except:
                    dob = None

            nationality = request.POST.get(f"nationality_{p}", "").strip()
            passport = request.POST.get(f"passport_{p}", "").strip()

            if first_name and last_name and dob and nationality and passport:
                Passenger.objects.create(
                    booking=booking,
                    title=request.POST.get(f"title_{p}", "N/A"),
                    first_name=first_name,
                    middle_initial=middle_initial,
                    last_name=last_name,
                    dob=dob,
                    nationality=nationality,
                    passport=passport,
                    has_declaration=bool(request.POST.get(f"declaration_{p}")),
                    is_pwd=bool(request.POST.get(f"pwd_{p}")),
                )

        return redirect("Section_Detail", pk=section.id)

    return render(request, "Create/Create_Instruction.html", {
        "section": section,
        "routes": routes,
        "schedules": schedules,
    })