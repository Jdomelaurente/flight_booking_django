from django.http import HttpResponse
from django.template import loader
from django.shortcuts import redirect, render
from django.shortcuts import get_object_or_404, redirect
from datetime import datetime
from django.utils import timezone
from django.http import JsonResponse
from django.shortcuts import render

import hashlib
from .models import User
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from collections import defaultdict
from django.utils.timezone import localtime
from django.shortcuts import render
from .models import Schedule
from datetime import datetime
from django.utils import timezone
from datetime import datetime


from .models import Flight
from .models import Route
from .models import Schedule
from .models import Seat
from .models import Airport
from .models import Airline
from .models import Aircraft
from .models import SeatClass
from .models import Booking
from .models import BookingDetail
from .models import Payment
from .models import CheckInDetail
from .models import Student
from .models import PassengerInfo
from .models import  TrackLog

# ------------------------------------ Users ----------------------------------------------------------

# main
def main(request):
    template = loader.get_template('main.html')
    return HttpResponse(template.render())

def admin_dashboard(request):
    return render(request, "dashboard.html")

def instructor_dashboard(request):
    return render(request, "instructor_dashboard.html")


# profile
def profile_view(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login")  # redirect if not logged in

    user = User.objects.get(id=user_id)

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if username:
            user.username = username

        if password:
            user.set_password(password)  # hash the new password

        user.save()
        return redirect("profile")  # reload the profile page

    return render(request, "profile.html", {"user": user})

# Hash
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# REGISTER
def register_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password2 = request.POST.get("password2")
        role = request.POST.get("role")

        if password != password2:
            messages.error(request, "Passwords do not match.")
            return redirect("register")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("register")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect("register")

        User.objects.create(
            username=username,
            email=email,
            password=hash_password(password),
            role=role
        )
        messages.success(request, "Account created. Please login.")
        return redirect("login")

    return render(request, "register.html")

# LOGIN
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = hash_password(request.POST.get("password"))

        try:
            user = User.objects.get(username=username, password=password)
            request.session["user_id"] = user.id
            request.session["username"] = user.username
            request.session["role"] = user.role  # store role in session

            # redirect based on role
            if user.role == "admin":
                return redirect("admin_dashboard")
            elif user.role == "instructor":
                return redirect("instructor_dashboard")
            elif user.role == "student":
                return redirect("student_dashboard")
            else:
                return redirect("login")  # fallback
        except User.DoesNotExist:
            messages.error(request, "Invalid username or password.")
            return redirect("login")

    return render(request, "login.html")


# LOGOUT
def logout_view(request):
    request.session.flush()
    return redirect("login")

# ------------------------------------Dashboard ----------------------------------------------------------

def dashboard(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login")

    user = User.objects.get(id=user_id)
    if user.role == "admin":
        return redirect("admin_dashboard")
    elif user.role == "instructor":
        return redirect("instructor_dashboard")
    return redirect("login")

from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from django.http import JsonResponse
from datetime import datetime, timedelta

from .models import Schedule, Booking, Payment

def admin_dashboard(request):
    date_filter = request.GET.get("date")
    date_range = request.GET.get("range")
    flight_range = request.GET.get("flight_range")
    revenue_type = request.GET.get("type", "route")  # default: by airline

    schedules = Schedule.objects.all()
    current_time = timezone.now()
    today = current_time.date()

    # -------------------
    # Handle single date filter
    # -------------------
    if date_filter:
        try:
            selected_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
            schedules = schedules.filter(departure_time__date=selected_date)
        except ValueError:
            selected_date = None
    else:
        selected_date = None

    # -------------------
    # Flight Status Counters
    # -------------------
    schedule_data = []
    total_open = total_closed = total_on_flight = total_arrived = 0

    for s in schedules:
        if current_time < s.departure_time:
            status = s.status
            if status == "Open":
                total_open += 1
            else:
                total_closed += 1
        elif s.departure_time <= current_time < s.arrival_time:
            status = "On Flight"
            total_on_flight += 1
        else:
            status = "Arrived"
            total_arrived += 1

        schedule_data.append({"schedule": s, "status": status})

    # -------------------
    # Latest Bookings
    # -------------------
    recent_bookings = (
        Booking.objects
        .select_related("student")
        .prefetch_related("details__schedule__flight", "details__seat")
        .filter(status="Confirmed")
        .order_by("-created_at")[:10]
    )

    # -------------------
    # Extra Stats
    # -------------------
    passenger_today = Booking.objects.filter(created_at__date=today).count()
    total_bookings = Booking.objects.count()
    total_revenue = Payment.objects.aggregate(total=Sum("amount"))["total"] or 0

    # -------------------
    # Ticket Sales Chart
    # -------------------
    if date_range == "week":
        start = today - timedelta(days=today.weekday())
        days = [start + timedelta(days=i) for i in range(7)]
    elif date_range == "month":
        start = today.replace(day=1)
        days = [start + timedelta(days=i) for i in range((today - start).days + 1)]
    elif date_range == "year":
        months = [datetime(today.year, m, 1).date() for m in range(1, 13)]
        ticket_sales_data = [
            Booking.objects.filter(
                created_at__year=today.year,
                created_at__month=month.month,
                status="Confirmed"
            ).count() for month in months
        ]
        ticket_sales_labels = [month.strftime("%b") for month in months]
    else:
        # Default: last 7 days
        days = [today - timedelta(days=i) for i in range(6, -1, -1)]

    if date_range != "year":
        ticket_sales_data = [
            Booking.objects.filter(created_at__date=day, status="Confirmed").count()
            for day in days
        ]
        ticket_sales_labels = [day.strftime("%b %d") for day in days]

    total_tickets_sold = sum(ticket_sales_data)

    # -------------------
    # Flight Schedule Chart
    # -------------------
    if flight_range == "month":
        start = today.replace(day=1)
        days = [start + timedelta(days=i) for i in range((today - start).days + 1)]
        flight_schedule_data = [
            Schedule.objects.filter(departure_time__date=day).count() for day in days
        ]
        flight_schedule_labels = [day.strftime("%b %d") for day in days]

    elif flight_range == "year":
        months = [datetime(today.year, m, 1).date() for m in range(1, 13)]
        flight_schedule_data = [
            Schedule.objects.filter(
                departure_time__year=today.year,
                departure_time__month=month.month
            ).count() for month in months
        ]
        flight_schedule_labels = [month.strftime("%b") for month in months]

    else:
        # Default & "week"
        start = today - timedelta(days=today.weekday())
        days = [start + timedelta(days=i) for i in range(7)]
        flight_schedule_data = [
            Schedule.objects.filter(departure_time__date=day).count() for day in days
        ]
        flight_schedule_labels = [day.strftime("%a") for day in days]

    # -------------------
    # Revenue Breakdown
    # -------------------
    if revenue_type == "class":
        revenue_qs = (
            Payment.objects
            .values("booking__details__seat_class__name")
            .annotate(total=Sum("amount"))
        )
        revenue_labels = [r["booking__details__seat_class__name"] or "Unknown" for r in revenue_qs]
        revenue_data = [float(r["total"]) for r in revenue_qs]

    else:  # revenue by airline
        revenue_qs = (
            Payment.objects
            .values("booking__details__schedule__flight__airline__name")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )
        revenue_labels = [
            r["booking__details__schedule__flight__airline__name"] or "Unknown Airline"
            for r in revenue_qs
        ]
        revenue_data = [float(r["total"]) for r in revenue_qs]

    # -------------------
    # AJAX Response
    # -------------------
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        chart_type = request.GET.get("chart")
        if chart_type == "flight":
            return JsonResponse({"labels": flight_schedule_labels, "data": flight_schedule_data})
        elif chart_type == "revenue":
            return JsonResponse({"labels": revenue_labels, "data": revenue_data})
        else:
            return JsonResponse({"labels": ticket_sales_labels, "data": ticket_sales_data})

    # -------------------
    # Render full dashboard page
    # -------------------
    return render(request, "dashboard.html", {
        "schedule_data": schedule_data,
        "selected_date": selected_date,
        "username": request.session.get("username"),

        # Flight status counters
        "total_open": total_open,
        "total_closed": total_closed,
        "total_on_flight": total_on_flight,
        "total_arrived": total_arrived,

        # Extra stats
        "passenger_today": passenger_today,
        "total_bookings": total_bookings,
        "total_revenue": total_revenue,

        # Charts
        "ticket_sales_labels": ticket_sales_labels,
        "ticket_sales_data": ticket_sales_data,
        "total_tickets_sold": total_tickets_sold,
        "flight_schedule_labels": flight_schedule_labels,
        "flight_schedule_data": flight_schedule_data,

        # Revenue Breakdown
        "revenue_labels": revenue_labels,
        "revenue_data": revenue_data,

        "recent_bookings": recent_bookings,
    })



# ------------------------------------ ASSETS ----------------------------------------------------------


# ---------------------------
# add-ons
# ---------------------------

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import AddOn

def addon_view(request):
    addons = AddOn.objects.all()
    return render(request, 'asset/addon/addons.html', {'addons': addons})

def add_addon(request):
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST.get('description', '')
        price = request.POST['price']
        AddOn.objects.create(name=name, description=description, price=price)
        messages.success(request, 'Add-On added successfully!')
        return redirect('addon')
    return render(request, 'asset/addon/add_addon.html')

def update_addon(request, id):
    addon = get_object_or_404(AddOn, id=id)
    if request.method == 'POST':
        addon.name = request.POST['name']
        addon.description = request.POST.get('description', '')
        addon.price = request.POST['price']
        addon.save()
        messages.success(request, 'Add-On updated successfully!')
        return redirect('addon')
    return render(request, 'asset/addon/edit_addon.html', {'addon': addon})

def delete_addon(request, id):
    addon = get_object_or_404(AddOn, id=id)
    addon.delete()
    messages.success(request, 'Add-On deleted successfully!')
    return redirect('addon')

from django.shortcuts import redirect
from django.contrib import messages
from .models import AddOn
import openpyxl

# 📥 Import Add-Ons from Excel
def import_addons(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_addons = []
            already_in_db = set(AddOn.objects.values_list("name", flat=True))

            seen_names_in_pass2 = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, description, price = row
                if not name:
                    continue
                clean_name = str(name).strip()

                # Detect duplicates inside the Excel file
                if clean_name in seen_names_in_pass2:
                    duplicates_in_file.add(clean_name)
                    continue
                seen_names_in_pass2.add(clean_name)

                # Skip if already exists in DB
                if clean_name in already_in_db:
                    continue

                # Create new Add-On
                AddOn.objects.create(
                    name=clean_name,
                    description=description or "",
                    price=price or 0.00
                )
                new_addons.append(clean_name)

            # ✅ Display user-friendly messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate add-on names found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_addons:
                messages.success(
                    request,
                    f"Successfully added: {', '.join(new_addons)}"
                )
            elif not duplicates_in_file:
                messages.info(request, "No new add-ons to add.")

        except Exception as e:
            messages.error(request, f"Error importing add-ons: {e}")

    else:
        messages.error(request, "No file uploaded.")

    return redirect("addon")

# ---------------------------
# Seat Class
# ---------------------------

# List
def seat_class_view(request):
    seat_classes = SeatClass.objects.all()
    return render(request, "asset/seat_class/seat_class.html", {"seat_classes": seat_classes})
    
import openpyxl
from django.contrib import messages

def import_seat_classes(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_classes = []
            already_in_db = set(
                SeatClass.objects.values_list("name", flat=True)
            )

            seen_names_in_pass2 = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, price_multiplier = row
                if not name:
                    continue
                clean_name = name.strip()

                # Detect duplicates inside file
                if clean_name in seen_names_in_pass2:
                    duplicates_in_file.add(clean_name)
                    continue
                seen_names_in_pass2.add(clean_name)

                # Skip if already exists in DB
                if clean_name in already_in_db:
                    continue

                # Add new seat class
                SeatClass.objects.create(
                    name=clean_name,
                    price_multiplier=price_multiplier
                )
                new_classes.append(clean_name)

            # Show messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate names found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_classes:
                messages.success(
                    request,
                    f"Successfully added: {', '.join(new_classes)}"
                )
            elif not duplicates_in_file:
                messages.info(request, "No new seat classes to add.")

        except Exception as e:
            messages.error(request, f"Error importing seat classes: {e}")

    return redirect("seat_class")



# Add
def add_seat_class(request):
    if request.method == "POST":
        name = request.POST.get("name")
        price_multiplier = request.POST.get("price_multiplier")
        description = request.POST.get("description")  # optional

        SeatClass.objects.create(
            name=name,
            price_multiplier=price_multiplier
        )
        return redirect("seat_class")
    return render(request, "asset/seat_class/add_seat_class.html")

# Update
def update_seat_class(request, seat_class_id):
    seat_class = get_object_or_404(SeatClass, pk=seat_class_id)
    if request.method == "POST":
        seat_class.name = request.POST.get("name")
        seat_class.price_multiplier = request.POST.get("price_multiplier")
        seat_class.save()
        return redirect("seat_class")
    return render(request, "asset/seat_class/update_seat_class.html", {"seat_class": seat_class})


# Delete
def delete_seat_class(request, seat_class_id):
    seat_class = get_object_or_404(SeatClass, pk=seat_class_id)
    seat_class.delete()
    return redirect("seat_class")


# ---------------------------
# Aircraft
# ---------------------------

# List
def aircraft_view(request):
    aircrafts = Aircraft.objects.all()
    return render(request, 'asset/aircraft/aircraft.html', {"aircrafts": aircrafts})

import openpyxl
from django.contrib import messages

def import_aircrafts(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_aircrafts = []
            already_in_db = set(
                Aircraft.objects.values_list("model", "airline__name")
            )

            seen_aircrafts_in_file = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                model, capacity, airline_name = row
                if not model or not capacity or not airline_name:
                    continue

                clean_model = model.strip()
                clean_airline = airline_name.strip()
                key = (clean_model, clean_airline)

                # Detect duplicates inside file
                if key in seen_aircrafts_in_file:
                    duplicates_in_file.add(f"{clean_model} ({clean_airline})")
                    continue
                seen_aircrafts_in_file.add(key)

                # Skip if already exists in DB
                if key in already_in_db:
                    continue

                try:
                    airline = Airline.objects.get(name__iexact=clean_airline)
                except Airline.DoesNotExist:
                    continue  # skip if airline not found

                # Add new aircraft
                Aircraft.objects.create(
                    model=clean_model,
                    capacity=capacity,
                    airline=airline
                )
                new_aircrafts.append(f"{clean_model} ({airline.name})")

            # Show messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate aircraft found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_aircrafts:
                messages.success(
                    request,
                    f"Successfully added: {', '.join(new_aircrafts)}"
                )
            elif not duplicates_in_file:
                messages.info(request, "No new aircraft to add.")

        except Exception as e:
            messages.error(request, f"Error importing aircraft: {e}")

    return redirect("aircraft")



# Add
def add_aircraft(request):
    if request.method == "POST":
        model = request.POST.get("model")
        capacity = request.POST.get("capacity")
        airline_id = request.POST.get("airline")  # comes from dropdown
        airline = Airline.objects.get(id=airline_id)

        Aircraft.objects.create(
            model=model,
            capacity=capacity,
            airline=airline
        )
        return redirect("aircraft")

    airlines = Airline.objects.all()  # for dropdown
    return render(request, "asset/aircraft/add_aircraft.html", {"airlines": airlines})

# Update
def update_aircraft(request, aircraft_id):
    aircraft = get_object_or_404(Aircraft, id=aircraft_id)
    airlines = Airline.objects.all()

    if request.method == "POST":
        aircraft.model = request.POST["model"]
        aircraft.capacity = request.POST["capacity"]
        aircraft.type = request.POST["type"]
        airline_id = request.POST["airline"]
        aircraft.airline = get_object_or_404(Airline, id=airline_id)

        aircraft.save()
        return redirect("aircraft")

    return render(
        request,
        "asset/aircraft/update_aircraft.html",
        {"aircraft": aircraft, "airlines": airlines}
    )

# Delete
def delete_aircraft(request, aircraft_id):
    aircraft = get_object_or_404(Aircraft, id=aircraft_id)
    aircraft.delete()
    return redirect("aircraft")


# ---------------------------
# Airlines
# ---------------------------

# List
def airline_view(request):
    airlines = Airline.objects.all()
    return render(request, "asset/airline/airline.html", {"airlines": airlines})

import openpyxl
from django.contrib import messages

def import_airlines(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_airlines = []
            already_in_db_codes = set(Airline.objects.values_list("code", flat=True))
            already_in_db_names = set(Airline.objects.values_list("name", flat=True))

            seen_codes_in_file = set()
            seen_names_in_file = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                code, name = row
                if not code or not name:
                    continue

                clean_code = code.strip()
                clean_name = name.strip()

                # Detect duplicates inside file (by code or name)
                if (clean_code in seen_codes_in_file) or (clean_name.lower() in seen_names_in_file):
                    duplicates_in_file.add(f"{clean_code} - {clean_name}")
                    continue
                seen_codes_in_file.add(clean_code)
                seen_names_in_file.add(clean_name.lower())

                # Skip if already exists in DB (by code or name)
                if clean_code in already_in_db_codes or clean_name in already_in_db_names:
                    continue

                # Add new airline
                Airline.objects.create(
                    code=clean_code,
                    name=clean_name
                )
                new_airlines.append(f"{clean_code} - {clean_name}")

            # Show messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate airlines found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_airlines:
                messages.success(
                    request,
                    f"Successfully added: {', '.join(new_airlines)}"
                )
            elif not duplicates_in_file:
                messages.info(request, "No new airlines to add.")

        except Exception as e:
            messages.error(request, f"Error importing airlines: {e}")

    return redirect("airline")


# Add
def add_airline(request):
    if request.method == "POST":
        name = request.POST.get("name")
        code = request.POST.get("code")
        Airline.objects.create(name=name, code=code)
        return redirect("airline")
    return render(request, "asset/airline/add_airline.html")

# Update
def update_airline(request, airline_id):
    airline = get_object_or_404(Airline, id=airline_id)
    if request.method == "POST":
        airline.name = request.POST.get("name")
        airline.code = request.POST.get("code")
        airline.save()
        return redirect("airline")
    return render(request, "asset/airline/update_airline.html", {"airline": airline})

# Delete
def delete_airline(request, airline_id):
    airline = get_object_or_404(Airline, id=airline_id)
    airline.delete()
    return redirect("airline")

# ---------------------------
# Airport
# ---------------------------

# List
def airport_view(request):
    airports = Airport.objects.all()
    return render(request, "asset/airport/airport.html", {"airports": airports})

import openpyxl
from django.contrib import messages

def import_airports(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_airports = []
            already_in_db_codes = set(
                Airport.objects.values_list("code", flat=True)
            )

            seen_codes_in_file = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                code, name, location = row
                if not code or not name:
                    continue

                clean_code = code.strip().upper()
                clean_name = name.strip()
                clean_location = (location or "").strip()

                # Detect duplicates inside file
                if clean_code in seen_codes_in_file:
                    duplicates_in_file.add(clean_code)
                    continue
                seen_codes_in_file.add(clean_code)

                # Skip if already exists in DB (unique by code)
                if clean_code in already_in_db_codes:
                    continue

                # Add new airport
                Airport.objects.create(
                    code=clean_code,
                    name=clean_name,
                    location=clean_location,
                )
                new_airports.append(f"{clean_code} - {clean_name}")

            # Show messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate airport codes found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_airports:
                messages.success(
                    request,
                    f"Successfully added: {', '.join(new_airports)}"
                )
            elif not duplicates_in_file:
                messages.info(request, "No new airports to add.")

        except Exception as e:
            messages.error(request, f"Error importing airports: {e}")

    return redirect("airport")



from django.db import IntegrityError
# Add
def add_airport(request):
    error_message = None

    if request.method == "POST":
        name = request.POST.get("name")
        code = request.POST.get("code")
        location = request.POST.get("location")

        # Check for duplicate airport code
        if Airport.objects.filter(code=code).exists():
            error_message = f"Airport with code {code} already exists."
        else:
            try:
                Airport.objects.create(
                    name=name,
                    code=code,
                    location=location,
                )
                return redirect("airport")
            except IntegrityError:
                error_message = "Something went wrong while saving."

    return render(request, "asset/airport/add_airport.html", {"error_message": error_message})


# Update
def update_airport(request, airport_id):
    airport = get_object_or_404(Airport, pk=airport_id)

    if request.method == "POST":
        airport.name = request.POST.get("name")
        airport.code = request.POST.get("code")
        airport.location = request.POST.get("location")
        airport.save()
        return redirect("airport")

    return render(request, "asset/airport/update_airport.html", {"airport": airport})

# Delete
def delete_airport(request, airport_id):
    airport = get_object_or_404(Airport, pk=airport_id)
    airport.delete()
    return redirect("airport")

# ------------------------------------------ Manage Flight ----------------------------------------

# ---------------------------
# Flight
# ---------------------------

# List
def flight_view(request):
    flights = Flight.objects.all()
    airlines = Airline.objects.all()
    routes = Route.objects.all()
    return render(request, "manage_flight/flight/flight.html", {
        "flights": flights,
        "airlines": airlines,
        "routes": routes
    })

import pandas as pd
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist

def import_flights(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            new_flights = []
            errors = []

            # Already in DB (flight numbers)
            already_in_db = set(Flight.objects.values_list("flight_number", flat=True))

            seen_flight_numbers = set()

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                flight_number, airline_name, aircraft_model, route_str = row
                if not flight_number or not airline_name or not aircraft_model or not route_str:
                    continue

                clean_flight_number = str(flight_number).strip()

                # Check duplicate in file (only skip if it was already processed once)
                if clean_flight_number in seen_flight_numbers:
                    duplicates_in_file.add(clean_flight_number)
                    continue
                seen_flight_numbers.add(clean_flight_number)

                # Skip if already exists in DB
                if clean_flight_number in already_in_db:
                    continue

                # Validate airline
                try:
                    airline = Airline.objects.get(name__iexact=airline_name.strip())
                except Airline.DoesNotExist:
                    errors.append(f"Row {idx}: Airline '{airline_name}' not found")
                    continue

                # Validate aircraft
                try:
                    aircraft = Aircraft.objects.get(model__iexact=aircraft_model.strip(), airline=airline)
                except Aircraft.DoesNotExist:
                    errors.append(f"Row {idx}: Aircraft '{aircraft_model}' not found for airline '{airline_name}'")
                    continue

                # Parse route
                if "-" not in route_str:
                    errors.append(f"Row {idx}: Invalid route format '{route_str}'")
                    continue
                origin_code, dest_code = [x.strip().upper() for x in route_str.split("-", 1)]

                try:
                    route = Route.objects.get(
                        origin_airport__code=origin_code,
                        destination_airport__code=dest_code
                    )
                except Route.DoesNotExist:
                    errors.append(f"Row {idx}: Route {origin_code}-{dest_code} not found")
                    continue

                # Create flight
                Flight.objects.create(
                    flight_number=clean_flight_number,
                    airline=airline,
                    aircraft=aircraft,
                    route=route,
                )
                new_flights.append(clean_flight_number)  # ✅ mark as added

            # Show messages
            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate flight numbers in file (skipped): {', '.join(duplicates_in_file)}"
                )
            if new_flights:
                messages.success(
                    request,
                    f"Successfully added flights: {', '.join(new_flights)}"
                )
            if errors:
                for err in errors:
                    messages.error(request, err)
            elif not new_flights and not duplicates_in_file:
                messages.info(request, "No new flights to add.")

        except Exception as e:
            messages.error(request, f"Error importing flights: {e}")

    return redirect("flight")


# Add 
def add_flight(request):
    if request.method == "POST":
        flight_number = request.POST.get("flight_number")
        airline_id = request.POST.get("airline")
        aircraft_id = request.POST.get("aircraft")
        route_id = request.POST.get("route")

        airline = Airline.objects.get(id=airline_id)
        aircraft = Aircraft.objects.get(id=aircraft_id)
        route = Route.objects.get(id=route_id)

        Flight.objects.create(
            flight_number=flight_number,
            airline=airline,
            aircraft=aircraft,
            route=route,
        )
        return redirect("flight")

    # ⬇️ not used anymore since modal form is inside flight.html
    return redirect("flight")



# AJAX endpoint to load aircraft based on airline
def load_aircrafts(request):
    airline_id = request.GET.get("airline")
    aircrafts = Aircraft.objects.filter(airline_id=airline_id).values("id", "model")
    return JsonResponse(list(aircrafts), safe=False)

# Update
def update_flight(request, flight_id):
    flight = get_object_or_404(Flight, pk=flight_id)
    airlines = Airline.objects.all()
    aircrafts = Aircraft.objects.all()
    routes = Route.objects.all()

    if request.method == "POST":
        flight.flight_number = request.POST.get("flight_number")
        flight.airline = get_object_or_404(Airline, pk=request.POST.get("airline"))
        flight.aircraft = get_object_or_404(Aircraft, pk=request.POST.get("aircraft"))
        flight.route = get_object_or_404(Route, pk=request.POST.get("route"))
        flight.save()
        return redirect("flight")

    return render(request, "manage_flight/flight/update_flight.html", {
        "flight": flight,
        "airlines": airlines,
        "aircrafts": aircrafts,
        "routes": routes
    })

# Delete flight
from django.views.decorators.http import require_POST

@require_POST
def delete_flight(request, flight_id):
    flight = get_object_or_404(Flight, pk=flight_id)
    flight.delete()
    return redirect("flight")



# ---------------------------
# Route
# ---------------------------

# List
def route_view(request):
    routes = Route.objects.all()
    return render(request, "manage_flight/route/route.html", {"routes": routes})

def import_routes(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            duplicates_in_file = set()
            invalid_routes = set()
            new_routes = []

            # Collect already existing routes in DB as a set of keys "ORIGIN-DEST"
            already_in_db = set(
                f"{r.origin_airport.code.upper()}-{r.destination_airport.code.upper()}"
                for r in Route.objects.select_related("origin_airport", "destination_airport")
            )

            seen_routes_in_file = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                origin_code, destination_code, base_price = row
                if not origin_code or not destination_code:
                    continue

                clean_origin = origin_code.strip().upper()
                clean_destination = destination_code.strip().upper()
                route_key = f"{clean_origin}-{clean_destination}"

                # 🚫 Prevent routes with same origin and destination
                if clean_origin == clean_destination:
                    invalid_routes.add(route_key)
                    continue

                # Detect duplicates inside file
                if route_key in seen_routes_in_file:
                    duplicates_in_file.add(route_key)
                    continue
                seen_routes_in_file.add(route_key)

                # Skip if already exists in DB
                if route_key in already_in_db:
                    continue

                try:
                    origin = Airport.objects.get(code__iexact=clean_origin)
                    destination = Airport.objects.get(code__iexact=clean_destination)
                except Airport.DoesNotExist:
                    continue  # Skip if airport not found in DB

                # Add new route
                Route.objects.create(
                    origin_airport=origin,
                    destination_airport=destination,
                    base_price=base_price or 0.00,
                )
                new_routes.append(route_key)

            # Show messages
            if invalid_routes:
                messages.warning(
                    request,
                    f"Invalid routes skipped (same origin & destination): {', '.join(invalid_routes)}"
                )

            if duplicates_in_file:
                messages.warning(
                    request,
                    f"Duplicate routes found in the file (skipped): {', '.join(duplicates_in_file)}"
                )

            if new_routes:
                messages.success(
                    request,
                    f"Successfully added routes: {', '.join(new_routes)}"
                )
            elif not duplicates_in_file and not invalid_routes:
                messages.info(request, "No new routes to add.")

        except Exception as e:
            messages.error(request, f"Error importing routes: {e}")

    return redirect("route")



# Add 
def add_route(request):
    airports = Airport.objects.all()
    if request.method == "POST":
        origin_id = request.POST.get("origin_airport")
        destination_id = request.POST.get("destination_airport")
        base_price = request.POST.get("base_price", 0.00)

        origin = get_object_or_404(Airport, pk=origin_id)
        destination = get_object_or_404(Airport, pk=destination_id)

        Route.objects.create(
            origin_airport=origin,
            destination_airport=destination,
            base_price=base_price
        )
        return redirect("route")

    return render(request, "manage_flight/route/add_route.html", {"airports": airports, "filtered_destinations": airports})



# Update 
def update_route(request, route_id):
    route = get_object_or_404(Route, pk=route_id)
    airports = Airport.objects.all()
    if request.method == "POST":
        route.origin_airport = get_object_or_404(Airport, pk=request.POST.get("origin_airport"))
        route.destination_airport = get_object_or_404(Airport, pk=request.POST.get("destination_airport"))
        route.base_price = request.POST.get("base_price", route.base_price)
        route.save()
        return redirect("route")
    return render(request, "manage_flight/route/update_route.html", {"route": route, "airports": airports})


# Delete 
def delete_route(request, route_id):
    route = get_object_or_404(Route, pk=route_id)
    route.delete()
    return redirect("route")


from django.utils import timezone

# ---------------------------
# Schedule
# ---------------------------


def schedule_view(request):
    schedules = Schedule.objects.all()

    schedule_data = []
    total_open = total_closed = total_on_flight = total_arrived = 0

    for s in schedules:
        s.update_status()  # 🔹 model decides correct status

        if s.status == "Open":
            total_open += 1
        elif s.status == "Closed":
            total_closed += 1
        elif s.status == "On Flight":
            total_on_flight += 1
        elif s.status == "Arrived":
            total_arrived += 1

        schedule_data.append({
            "schedule": s,
            "status": s.status
        })

    context = {
        "schedule_data": schedule_data,
        "total_open": total_open,
        "total_closed": total_closed,
        "total_on_flight": total_on_flight,
        "total_arrived": total_arrived,
    }
    return render(request, "manage_flight/schedule/schedule.html", context)

import openpyxl
from django.contrib import messages
from django.utils.dateparse import parse_datetime
from datetime import timedelta
from django.shortcuts import redirect
from .models import Flight, Schedule

def import_schedules(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        errors = []
        new_schedules = []
        allowance = timedelta(minutes=15)  # 15-min conflict allowance

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            seen_in_file = set()  # to detect duplicate rows inside Excel

            # Expected Excel format:
            # Flight Number | Departure Time | Arrival Time | Price
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                flight_number, departure_time, arrival_time, price = row

                if not flight_number or not departure_time or not arrival_time:
                    errors.append(f"Row {idx}: Missing required data")
                    continue

                # Get flight
                try:
                    flight = Flight.objects.get(flight_number=str(flight_number).strip())
                except Flight.DoesNotExist:
                    errors.append(f"Row {idx}: Flight '{flight_number}' not found")
                    continue

                # Parse datetimes properly
                dep_time = departure_time if isinstance(departure_time, datetime) else parse_datetime(str(departure_time))
                arr_time = arrival_time if isinstance(arrival_time, datetime) else parse_datetime(str(arrival_time))
                if not dep_time or not arr_time:
                    errors.append(f"Row {idx}: Invalid datetime format")
                    continue

                if dep_time >= arr_time:
                    errors.append(f"Row {idx}: Departure time must be before arrival time")
                    continue

                # Check if this exact row already appeared in the same Excel
                row_key = (flight.id, dep_time, arr_time)
                if row_key in seen_in_file:
                    errors.append(f"Row {idx}: Duplicate schedule in file (skipped)")
                    continue
                seen_in_file.add(row_key)

                origin = flight.route.origin_airport
                destination = flight.route.destination_airport

                # 🔹 Departure conflict check (same origin, within 15 mins)
                conflict_dep = Schedule.objects.filter(
                    flight__route__origin_airport=origin
                ).exclude(
                    flight=flight,
                    departure_time=dep_time,
                    arrival_time=arr_time
                ).filter(
                    departure_time__gte=dep_time - allowance,
                    departure_time__lte=dep_time + allowance
                ).exists()

                # 🔹 Arrival conflict check (same destination, within 15 mins)
                conflict_arr = Schedule.objects.filter(
                    flight__route__destination_airport=destination
                ).exclude(
                    flight=flight,
                    departure_time=dep_time,
                    arrival_time=arr_time
                ).filter(
                    arrival_time__gte=arr_time - allowance,
                    arrival_time__lte=arr_time + allowance
                ).exists()

                if conflict_dep:
                    errors.append(
                        f"Row {idx}: Another flight is departing from {origin.code} within 15 minutes."
                    )
                    continue

                if conflict_arr:
                    errors.append(
                        f"Row {idx}: Another flight is arriving at {destination.code} within 15 minutes."
                    )
                    continue

                # Avoid duplicates (same flight, same dep+arr in DB)
                if Schedule.objects.filter(
                    flight=flight,
                    departure_time=dep_time,
                    arrival_time=arr_time
                ).exists():
                    errors.append(f"Row {idx}: Schedule already exists in database")
                    continue

                # Create schedule
                Schedule.objects.create(
                    flight=flight,
                    departure_time=dep_time,
                    arrival_time=arr_time,
                    price=price or flight.route.base_price,
                    status="Open"
                )
                new_schedules.append(f"{flight_number} ({dep_time.strftime('%Y-%m-%d %H:%M')})")

            # Messages
            if new_schedules:
                messages.success(request, f"Successfully added schedules: {', '.join(new_schedules)}")
            for err in errors:
                messages.error(request, err)
            if not new_schedules and not errors:
                messages.info(request, "No new schedules to add.")

        except Exception as e:
            messages.error(request, f"Error reading file: {e}")

    return redirect("schedule")



from django.utils.dateparse import parse_datetime
from datetime import timedelta
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Flight, Schedule

def add_schedule(request):
    flights = Flight.objects.all()  # for the dropdown

    if request.method == "POST":
        flight_id = request.POST.get("flight")
        flight = get_object_or_404(Flight, id=flight_id)

        departure_time = request.POST.get("departure_time")
        arrival_time = request.POST.get("arrival_time")
        price = request.POST.get("price")

        # Convert to datetime
        dep_time = parse_datetime(departure_time)
        arr_time = parse_datetime(arrival_time)

        if not dep_time or not arr_time:
            messages.error(request, "Invalid date/time format.")
            return redirect("add_schedule")

        # Ensure departure is before arrival
        if dep_time >= arr_time:
            messages.error(request, "Departure time must be before arrival time.")
            return redirect("add_schedule")

        # Airports
        origin = flight.route.origin_airport
        destination = flight.route.destination_airport

        # 15-min allowance
        allowance = timedelta(minutes=15)

        # Conflict: departure at same origin within ±15 mins
        conflict_dep = Schedule.objects.filter(
            flight__route__origin_airport=origin,
            departure_time__range=(dep_time - allowance, dep_time + allowance)
        ).exists()

        # Conflict: arrival at same destination within ±15 mins
        conflict_arr = Schedule.objects.filter(
            flight__route__destination_airport=destination,
            arrival_time__range=(arr_time - allowance, arr_time + allowance)
        ).exists()

        if conflict_dep:
            messages.error(
                request, f"Another flight is departing from {origin.code} within 15 minutes."
            )
            return redirect("add_schedule")

        if conflict_arr:
            messages.error(
                request, f"Another flight is arriving at {destination.code} within 15 minutes."
            )
            return redirect("add_schedule")

        # ✅ Save schedule
        Schedule.objects.create(
            flight=flight,
            departure_time=dep_time,
            arrival_time=arr_time,
            price=price or flight.route.base_price,
            status="Open"
        )
        messages.success(request, "Schedule added successfully.")
        return redirect("schedule")

    return render(
        request,
        "manage_flight/schedule/add_schedule.html",
        {"flights": flights}
    )




# Update schedule
def update_schedule(request, schedule_id):
    schedule = get_object_or_404(Schedule, pk=schedule_id)
    flights = Flight.objects.all()
    if request.method == "POST":
        schedule.flight = get_object_or_404(Flight, pk=request.POST.get("flight"))
        schedule.departure_time = request.POST.get("departure_time")
        schedule.arrival_time = request.POST.get("arrival_time")
        schedule.save()
        return redirect("schedule")
    return render(request, "manage_flight/schedule/update_schedule.html", {"schedule": schedule, "flights": flights})


# Delete schedule
def delete_schedule(request, schedule_id):
    schedule = get_object_or_404(Schedule, pk=schedule_id)
    schedule.delete()
    return redirect("schedule")


# -----------------------------
# Seat
# -----------------------------

# List


def seat_view(request):
    seats = Seat.objects.all()
    flights = Flight.objects.all()

    flight_id = request.GET.get("flight")
    if flight_id:
        seats = seats.filter(schedule__flight__id=flight_id)

    return render(request, "manage_flight/seat/seat.html", {"seats": seats, "flights": flights})


# Add
def add_seat(request):
    schedules = Schedule.objects.all()
    seat_classes = SeatClass.objects.all()
    
    if request.method == "POST":
        schedule_id = request.POST.get("schedule")
        seat_class_id = request.POST.get("seat_class")
        seat_number = request.POST.get("seat_number")
        is_available = "is_available" in request.POST

        schedule = get_object_or_404(Schedule, id=schedule_id)
        seat_class = get_object_or_404(SeatClass, id=seat_class_id)

        Seat.objects.create(
            schedule=schedule,
            seat_class=seat_class,
            seat_number=seat_number,
            is_available=is_available
        )
        return redirect("seat")
    
    return render(request, "manage_flight/seat/add_seat.html", {"schedules": schedules, "seat_classes": seat_classes})

# Update 
def update_seat(request, seat_id):
    seat = get_object_or_404(Seat, id=seat_id)
    schedules = Schedule.objects.all()
    seat_classes = SeatClass.objects.all()
    
    if request.method == "POST":
        schedule_id = request.POST.get("schedule")
        seat_class_id = request.POST.get("seat_class")
        seat_number = request.POST.get("seat_number")
        is_available = "is_available" in request.POST

        seat.schedule = get_object_or_404(Schedule, id=schedule_id)
        seat.seat_class = get_object_or_404(SeatClass, id=seat_class_id)
        seat.seat_number = seat_number
        seat.is_available = is_available
        seat.save()
        return redirect("seat")
    
    return render(request, "manage_flight/seat/update_seat.html", {
        "seat": seat,
        "schedules": schedules,
        "seat_classes": seat_classes
    })

# Delete
def delete_seat(request, seat_id):
    seat = get_object_or_404(Seat, id=seat_id)
    seat.delete()
    return redirect("seat")

# ---------------------------------------Booking Info-----------------------------------------------

# ---------------------------
# Booking
# ---------------------------

# List 
from django.utils import timezone
from datetime import timedelta
from .models import Booking, BookingDetail, Seat

def booking_view(request):
    # Auto-delete pending bookings older than 15 minutes
    cutoff_time = timezone.now() - timedelta(minutes=10)
    expired_bookings = Booking.objects.filter(status="Pending", created_at__lt=cutoff_time)

    for booking in expired_bookings:
        # Free all seats in booking details
        for detail in booking.details.all():  # use related_name="details"
            if detail.seat:
                detail.seat.is_available = True
                detail.seat.save()

        # Delete booking
        booking.delete()

    bookings = Booking.objects.all().order_by('-created_at')
    return render(request, "booking_info/booking/booking.html", {"bookings": bookings})




from datetime import date

def calculate_booking_price(schedule, seat):
    # 1. Route Base Price
    route = schedule.flight.route
    base_price = route.base_price  

    # 2. Seat Multiplier
    seat_multiplier = seat.seat_class.price_multiplier if seat and seat.seat_class else 1.0

    # 3. Days Before Flight → Booking Factor
    days_diff = (schedule.departure_time.date() - date.today()).days
    if days_diff >= 30:
        booking_factor = 0.8
    elif 7 <= days_diff < 30:
        booking_factor = 1.0
    else:  # 1–6 days
        booking_factor = 1.5

    # Final Price
    return round(base_price * float(seat_multiplier) * booking_factor, 2)

# add
from django.shortcuts import render, redirect, get_object_or_404
from .models import Booking, BookingDetail, Student, Schedule, Seat, PassengerInfo

def add_booking(request):
    students = Student.objects.all()
    schedules = Schedule.objects.filter(status="Open").order_by("departure_time")

    if request.method == "POST":
        student_id = request.POST.get("student")
        trip_type = request.POST.get("trip_type")
        status = request.POST.get("status", "Pending")
        student = get_object_or_404(Student, id=student_id)

        booking = Booking.objects.create(
            student=student,
            trip_type=trip_type,
            status=status
        )

        # ---------------- Collect Passengers ----------------
        passenger_keys = [key for key in request.POST if key.startswith("passenger_first_name_")]
        passenger_count = len(passenger_keys)
        passengers = {}

        for i in range(1, passenger_count + 1):
            p_type = request.POST.get(f"passenger_type_{i}", "adult")
            linked_adult_index = request.POST.get(f"linked_adult_{i}")

            passenger = PassengerInfo.objects.create(
                first_name=request.POST.get(f"passenger_first_name_{i}"),
                middle_name=request.POST.get(f"passenger_middle_name_{i}", ""),
                last_name=request.POST.get(f"passenger_last_name_{i}"),
                gender=request.POST.get(f"passenger_gender_{i}", "N/A"),
                date_of_birth=request.POST.get(f"passenger_dob_{i}"),
                passenger_type=p_type
            )
            passengers[str(i)] = {
                "obj": passenger,
                "linked_adult_index": linked_adult_index,
                "type": p_type
            }

        # ---------------- Outbound ----------------
        if trip_type in ["one_way", "round_trip"]:
            outbound_schedule_id = request.POST.get("outbound_schedule")
            if outbound_schedule_id:
                outbound_schedule = get_object_or_404(Schedule, id=outbound_schedule_id)

                for idx in range(1, passenger_count + 1):
                    p = passengers[str(idx)]
                    seat = None

                    if p["type"] == "infant" and p["linked_adult_index"]:
                        # Infant: use adult's seat
                        adult_seat_id = request.POST.get(f"outbound_seat_{p['linked_adult_index']}")
                        seat = Seat.objects.filter(id=adult_seat_id).first() if adult_seat_id else None
                    else:
                        # Adult/child: normal seat
                        seat_id = request.POST.get(f"outbound_seat_{idx}")
                        seat = Seat.objects.filter(id=seat_id).first() if seat_id else None

                    BookingDetail.objects.create(
                        booking=booking,
                        passenger=p["obj"],
                        schedule=outbound_schedule,
                        seat=seat,
                        seat_class=seat.seat_class if seat else None
                    )

        # ---------------- Return ----------------
        if trip_type == "round_trip":
            return_schedule_id = request.POST.get("return_schedule")
            if return_schedule_id:
                return_schedule = get_object_or_404(Schedule, id=return_schedule_id)

                for idx in range(1, passenger_count + 1):
                    p = passengers[str(idx)]
                    seat = None

                    if p["type"] == "infant" and p["linked_adult_index"]:
                        adult_seat_id = request.POST.get(f"return_seat_{p['linked_adult_index']}")
                        seat = Seat.objects.filter(id=adult_seat_id).first() if adult_seat_id else None
                    else:
                        seat_id = request.POST.get(f"return_seat_{idx}")
                        seat = Seat.objects.filter(id=seat_id).first() if seat_id else None

                    BookingDetail.objects.create(
                        booking=booking,
                        passenger=p["obj"],
                        schedule=return_schedule,
                        seat=seat,
                        seat_class=seat.seat_class if seat else None
                    )

        # ---------------- Multi-City ----------------
        if trip_type == "multi_city":
            for key in request.POST:
                if key.startswith("multi_schedule_"):
                    leg_id = key.split("_")[2]
                    schedule_id = request.POST.get(f"multi_schedule_{leg_id}")

                    if schedule_id:
                        schedule = get_object_or_404(Schedule, id=schedule_id)

                        for idx in range(1, passenger_count + 1):
                            p = passengers[str(idx)]
                            seat = None

                            if p["type"] == "infant" and p["linked_adult_index"]:
                                adult_seat_id = request.POST.get(f"multi_seat_{leg_id}_{p['linked_adult_index']}")
                                seat = Seat.objects.filter(id=adult_seat_id).first() if adult_seat_id else None
                            else:
                                seat_id = request.POST.get(f"multi_seat_{leg_id}_{idx}")
                                seat = Seat.objects.filter(id=seat_id).first() if seat_id else None

                            BookingDetail.objects.create(
                                booking=booking,
                                passenger=p["obj"],
                                schedule=schedule,
                                seat=seat,
                                seat_class=seat.seat_class if seat else None
                            )

        return redirect("booking")

    context = {
        "students": students,
        "schedules": schedules
    }
    return render(request, "booking_info/booking/add_booking.html", context)



from django.http import JsonResponse
from .models import Seat, Schedule

# 🔹 Fetch seats for a schedule (outbound or return)
def get_seats_for_schedule(request, schedule_id):
    try:
        schedule = Schedule.objects.get(id=schedule_id)
    except Schedule.DoesNotExist:
        return JsonResponse([], safe=False)

    # Only seats that belong to this schedule
    seats = Seat.objects.filter(schedule=schedule).select_related("seat_class")

    data = []
    for seat in seats:
        data.append({
            "id": seat.id,
            "label": f"{seat.seat_number} - {seat.seat_class.name} - {'Available' if seat.is_available else 'Booked'}",
            "is_available": seat.is_available
        })

    return JsonResponse(data, safe=False)

def get_return_schedules(request, outbound_id):
    try:
        outbound = Schedule.objects.get(id=outbound_id)
        origin = outbound.flight.route.origin_airport
        destination = outbound.flight.route.destination_airport
    except Schedule.DoesNotExist:
        return JsonResponse([], safe=False)

    return_schedules = Schedule.objects.filter(
        flight__route__origin_airport=destination,
        flight__route__destination_airport=origin,
        status="Open"
    ).order_by("departure_time")

    data = []
    for s in return_schedules:
        if s.id != outbound.id:
            data.append({
                "id": s.id,
                "flight_number": s.flight.flight_number,
                "origin": s.flight.route.origin_airport.code,
                "destination": s.flight.route.destination_airport.code,
                "departure_time": s.departure_time.strftime("%Y-%m-%d %H:%M")
            })

    return JsonResponse(data, safe=False)
    


# Update
def update_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    students = Student.objects.all()
    schedules = Schedule.objects.all()
    seats = Seat.objects.filter(is_available=True) | Seat.objects.filter(id=booking.seat.id if booking.seat else None)

    if request.method == "POST":
        student_id = request.POST.get("student")
        schedule_id = request.POST.get("schedule")
        seat_id = request.POST.get("seat")
        status = request.POST.get("status")

        # Free previous seat if changed
        if booking.seat and (not seat_id or int(seat_id) != booking.seat.id):
            booking.seat.is_available = True
            booking.seat.save()

        booking.student = get_object_or_404(Student, id=student_id)
        booking.schedule = get_object_or_404(Schedule, id=schedule_id)
        booking.seat = get_object_or_404(Seat, id=seat_id) if seat_id else None
        booking.status = status
        booking.save()

        # Mark new seat as unavailable
        if booking.seat:
            booking.seat.is_available = False
            booking.seat.save()

        return redirect("booking")
    
    return render(request, "booking_info/booking/update_booking.html", {
        "booking": booking,
        "students": students,
        "schedules": schedules,
        "seats": seats,
    })

# Delete 
def delete_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    if booking.seat:
        booking.seat.is_available = True
        booking.seat.save()
    booking.delete()
    return redirect("booking")

# ---------------------------
# Payment
# ---------------------------

# List
def payment_view(request):
    payments = Payment.objects.all()
    return render(request, "booking_info/payment/payment.html", {"payments": payments})

# Add
def add_payment(request):
    bookings = Booking.objects.all()  # to select a booking for payment
    if request.method == "POST":
        booking_id = request.POST.get("booking")
        amount = request.POST.get("amount")
        method = request.POST.get("method")
        status = request.POST.get("status")
        transaction_id = request.POST.get("transaction_id", "")

        booking = get_object_or_404(Booking, id=booking_id)

        Payment.objects.create(
            booking=booking,
            amount=amount,
            method=method,
            status=status,
            transaction_id=transaction_id
        )
        messages.success(request, "Payment added successfully!")
        return redirect("payment")

    return render(request, "booking_info/payment/add_payment.html", {"bookings": bookings})

# Update
def update_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)
    bookings = Booking.objects.all()
    if request.method == "POST":
        booking_id = request.POST.get("booking")
        payment.amount = request.POST.get("amount")
        payment.method = request.POST.get("method")
        payment.status = request.POST.get("status")
        payment.transaction_id = request.POST.get("transaction_id", "")
        payment.booking = get_object_or_404(Booking, id=booking_id)
        payment.save()
        messages.success(request, "Payment updated successfully!")
        return redirect("payment")

    return render(request, "booking_info/payment/update_payment.html", {"payment": payment, "bookings": bookings})

# Delete 
def delete_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)
    payment.delete()
    messages.success(request, "Payment deleted successfully!")
    return redirect("payment")

# ---------------------------
# Booking_detail
# ---------------------------

# List 
def booking_detail_view(request):
    details = BookingDetail.objects.all()
    return render(request, "booking_info/booking_detail/booking_detail.html", {"details": details})

# Add
def add_booking_detail(request):
    bookings = Booking.objects.all()
    flights = Flight.objects.all()
    seats = Seat.objects.filter(is_available=True)  # ✅ only show available seats
    seat_classes = SeatClass.objects.all()

    if request.method == "POST":
        booking_id = request.POST.get("booking")
        flight_id = request.POST.get("flight")
        seat_id = request.POST.get("seat")
        seat_class_id = request.POST.get("seat_class")

        booking = get_object_or_404(Booking, id=booking_id)
        flight = get_object_or_404(Flight, id=flight_id)
        seat = get_object_or_404(Seat, id=seat_id) if seat_id else None
        seat_class = get_object_or_404(SeatClass, id=seat_class_id) if seat_class_id else None

        # ✅ Create booking detail
        BookingDetail.objects.create(
            booking=booking,
            flight=flight,
            seat=seat,
            seat_class=seat_class
        )

        # ✅ Mark seat as unavailable
        if seat:
            seat.is_available = False
            seat.save()

        messages.success(request, "Booking detail added successfully!")
        return redirect("booking_detail")

    return render(request, "booking_info/booking_detail/add_booking_detail.html", {
        "bookings": bookings,
        "flights": flights,
        "seats": seats,
        "seat_classes": seat_classes
    })



# Update
def update_booking_detail(request, detail_id):
    detail = get_object_or_404(BookingDetail, id=detail_id)
    bookings = Booking.objects.all()
    flights = Flight.objects.all()
    seats = Seat.objects.filter(is_available=True) | Seat.objects.filter(id=detail.seat_id)
    seat_classes = SeatClass.objects.all()

    if request.method == "POST":
        detail.booking = get_object_or_404(Booking, id=request.POST.get("booking"))
        detail.flight = get_object_or_404(Flight, id=request.POST.get("flight"))

        # ✅ Seat logic
        old_seat = detail.seat
        new_seat_id = request.POST.get("seat")
        new_seat = get_object_or_404(Seat, id=new_seat_id) if new_seat_id else None

        if old_seat and old_seat != new_seat:
            old_seat.is_available = True
            old_seat.save()

        if new_seat and new_seat.is_available:
            new_seat.is_available = False
            new_seat.save()
            detail.seat = new_seat

        # ✅ Seat class
        seat_class_id = request.POST.get("seat_class")
        detail.seat_class = get_object_or_404(SeatClass, id=seat_class_id) if seat_class_id else None

        detail.save()
        messages.success(request, "Booking detail updated successfully!")
        return redirect("booking_detail")

    return render(request, "booking_info/booking_detail/update_booking_detail.html", {
        "detail": detail,
        "bookings": bookings,
        "flights": flights,
        "seats": seats,
        "seat_classes": seat_classes
    })
    

# Delete
def delete_booking_detail(request, detail_id):
    detail = get_object_or_404(BookingDetail, id=detail_id)

    # ✅ Restore seat availability
    if detail.seat:
        detail.seat.is_available = True
        detail.seat.save()

    detail.delete()
    messages.success(request, "Booking detail deleted successfully!")
    return redirect("booking_detail")


# ---------------------------------------Passenger Info---------------------------------------------

# ---------------------------
# Passenger
# ---------------------------

# List 
def passenger_view(request):
    passengers = PassengerInfo.objects.all()
    return render(request, "passenger_info/passenger/passenger.html", {"passengers": passengers})

# Add 
def add_passenger(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        middle_name = request.POST.get("middle_name")
        gender = request.POST.get("gender")
        date_of_birth = request.POST.get("date_of_birth")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        passport_number = request.POST.get("passport_number")

        PassengerInfo.objects.create(
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            gender=gender,
            date_of_birth=date_of_birth,
            phone=phone,
            email=email,
            passport_number=passport_number
        )
        messages.success(request, "Passenger added successfully!")
        return redirect("passenger")

    return render(request, "passenger_info/passenger/add_passenger.html")

# Update
def update_passenger(request, passenger_id):
    passenger = get_object_or_404(PassengerInfo, id=passenger_id)

    if request.method == "POST":
        passenger.first_name = request.POST.get("first_name")
        passenger.last_name = request.POST.get("last_name")
        passenger.middle_name = request.POST.get("middle_name")
        passenger.gender = request.POST.get("gender")
        passenger.date_of_birth = request.POST.get("date_of_birth")
        passenger.phone = request.POST.get("phone")
        passenger.email = request.POST.get("email")
        passenger.passport_number = request.POST.get("passport_number")
        passenger.save()

        messages.success(request, "Passenger updated successfully!")
        return redirect("passenger")

    return render(request, "passenger_info/passenger/update_passenger.html", {"passenger": passenger})

# Delete 
def delete_passenger(request, passsenger_id):
    passenger = get_object_or_404(PassengerInfo, id=passsenger_id)
    passenger.delete()
    messages.success(request, "Passenger deleted successfully!")
    return redirect("passenger")

# -----------------------------
# Students
# -----------------------------

# List 
def check_in_view(request):
    checkins = CheckInDetail.objects.select_related("booking_detail__booking").all()
    return render(request, "passenger_info/check_in/check_in.html", {"checkins": checkins})

import openpyxl
from django.contrib import messages
from django.shortcuts import redirect
from .models import Student

def import_students(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        errors = []
        imported_students = []

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                student_number, first_name, middle_initial, last_name, email, phone, password = row

                # Check required fields
                if not student_number or not first_name or not last_name or not email:
                    errors.append(f"Row {idx}: Missing required data (Student Number, First Name, Last Name, or Email)")
                    continue

                # Check for duplicates
                if Student.objects.filter(student_number=student_number).exists():
                    errors.append(f"Row {idx}: Student number '{student_number}' already exists")
                    continue
                if Student.objects.filter(email=email).exists():
                    errors.append(f"Row {idx}: Email '{email}' already exists")
                    continue

                # Create student
                Student.objects.create(
                    student_number=student_number,
                    first_name=first_name,
                    middle_initial=middle_initial or "",
                    last_name=last_name,
                    email=email,
                    phone=phone or "",
                    password=password or "12345"  # default password if empty
                )
                imported_students.append(student_number)

            # Messages
            if imported_students:
                messages.success(
                    request, f"Successfully imported students: {', '.join(imported_students)}"
                )
            for err in errors:
                messages.error(request, err)
            if not imported_students and not errors:
                messages.info(request, "No students were imported.")

        except Exception as e:
            messages.error(request, f"Error importing students: {e}")

    return redirect("student")


# Add Check-In
def add_checkin(request):
    booking_details = BookingDetail.objects.all()

    if request.method == "POST":
        booking_detail_id = request.POST.get("booking_detail")
        boarding_pass = request.POST.get("boarding_pass")
        baggage_count = request.POST.get("baggage_count") or 0
        baggage_weight = request.POST.get("baggage_weight") or 0.0

        CheckInDetail.objects.create(
            booking_detail_id=booking_detail_id,
            boarding_pass=boarding_pass,
            baggage_count=baggage_count,
            baggage_weight=baggage_weight
        )
        messages.success(request, "Check-in added successfully!")
        return redirect("check_in")

    return render(request, "passenger_info/check_in/add_check_in.html", {"booking_details": booking_details})

# Update
def update_checkin(request, checkin_id):
    checkin = get_object_or_404(CheckInDetail, id=checkin_id)
    booking_details = BookingDetail.objects.all()

    if request.method == "POST":
        checkin.booking_detail_id = request.POST.get("booking_detail")
        checkin.boarding_pass = request.POST.get("boarding_pass")
        checkin.baggage_count = request.POST.get("baggage_count") or 0
        checkin.baggage_weight = request.POST.get("baggage_weight") or 0.0
        checkin.save()

        messages.success(request, "Check-in updated successfully!")
        return redirect("check_in")

    return render(request, "passenger_info/check_in/update_check_in.html", {"checkin": checkin, "booking_details": booking_details})

# Delete 
def delete_checkin(request, checkin_id):
    checkin = get_object_or_404(CheckInDetail, id=checkin_id)
    checkin.delete()
    messages.success(request, "Check-in deleted successfully!")
    return redirect("check_in")

# ---------------------------------------Student Info-----------------------------------------------

# -----------------------------
# Students
# -----------------------------

# List Students
def student_view(request):
    students = Student.objects.all().order_by('student_number')
    return render(request, "student_info/student/student.html", {"students": students})

# Add
def add_student(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")  # get raw password
        phone = request.POST.get("phone")
        student_number = request.POST.get("student_number")

        if not password:
            messages.error(request, "Password is required.")
            return redirect("add_student")

        # Hash the password before saving
        hashed_password = make_password(password)

        Student.objects.create(
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=hashed_password,  # save hashed password
            phone=phone,
            student_number=student_number
        )
        messages.success(request, "Student added successfully.")
        return redirect("student")
    
    return render(request, "student_info/student/add_student.html")


# Update
def update_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == "POST":
        student.first_name = request.POST.get("first_name")
        student.last_name = request.POST.get("last_name")
        student.email = request.POST.get("email")
        student.phone = request.POST.get("phone")
        student.student_number = request.POST.get("student_number")
        password = request.POST.get("password")
        if password:
            student.password = password  # TODO: hash password
        student.save()
        messages.success(request, "Student updated successfully.")
        return redirect("student")
    
    return render(request, "student_info/student/update_student.html", {"student": student})

# Delete
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.delete()
    messages.success(request, "Student deleted successfully.")
    return redirect("student")

# -----------------------------
# Tracklog
# -----------------------------

# List TrackLogs
def tracklog_view(request):
    tracklogs = TrackLog.objects.select_related("student").all()
    return render(request, "student_info/track_log/tracklog.html", {"tracklogs": tracklogs})

# Add TrackLog
def add_tracklog(request):
    students = Student.objects.all()

    if request.method == "POST":
        student_id = request.POST.get("student")
        action = request.POST.get("action")

        TrackLog.objects.create(
            student_id=student_id,
            action=action,
        )
        messages.success(request, "TrackLog added successfully!")
        return redirect("tracklog")

    return render(request, "student_info/track_log/add_tracklog.html", {"students": students})

# Update TrackLog
def update_tracklog(request, tracklog_id):
    tracklog = get_object_or_404(TrackLog, id=tracklog_id)
    students = Student.objects.all()

    if request.method == "POST":
        tracklog.student_id = request.POST.get("student")
        tracklog.action = request.POST.get("action")
        tracklog.save()

        messages.success(request, "TrackLog updated successfully!")
        return redirect("tracklog")

    return render(request, "student_info/track_log/update_tracklog.html", {"tracklog": tracklog, "students": students})

# Delete TrackLog
def delete_tracklog(request, tracklog_id):
    tracklog = get_object_or_404(TrackLog, id=tracklog_id)
    tracklog.delete()
    messages.success(request, "TrackLog deleted successfully!")
    return redirect("tracklog")

# --------------------------------sa payment og email ni -----------------

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.core.mail import EmailMessage
from django.conf import settings
from .models import Booking, Payment, BookingDetail, SeatClass
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Booking, Payment

from datetime import date

from datetime import date
from decimal import Decimal
from django.utils import timezone

def payment_success(request, booking_id):
    booking = get_object_or_404(Booking, pk=booking_id)
    payment = Payment.objects.filter(booking=booking).last()

    # Create payment if it doesn't exist
    if not payment:
        payment = Payment.objects.create(
            booking=booking,
            amount=booking.price,
            payment_method="Manual",
            status="Paid"
        )

    booking.status = "Confirmed"
    booking.save()

    # ----------------------------
    # Calculation for one-way / outbound
    # ----------------------------
    if booking.outbound_schedule:
        flight_schedule_date = booking.outbound_schedule.departure_time.date()
        route_base_price = booking.outbound_schedule.route.base_price
        seat_name = booking.outbound_seat.seat_class.name if booking.outbound_seat else "Economy"
        seat_multiplier = booking.outbound_seat.seat_class.price_multiplier if booking.outbound_seat else Decimal("1.0")
        days_diff = (flight_schedule_date - booking.created_at.date()).days

        # Determine booking factor
        if days_diff >= 30:
            booking_factor = Decimal("0.8")
        elif 7 <= days_diff <= 29:
            booking_factor = Decimal("1.0")
        else:
            booking_factor = Decimal("1.5")

        final_price = route_base_price * seat_multiplier * booking_factor
    else:
        flight_schedule_date = None
        route_base_price = 0
        seat_name = "Economy"
        seat_multiplier = 1.0
        days_diff = 0
        booking_factor = 1.0
        final_price = 0

    context = {
        "booking": booking,
        "payment": payment,
        "calc": {
            "booking_date": booking.created_at.date(),
            "flight_schedule": flight_schedule_date,
            "days_diff": days_diff,
            "route_base_price": route_base_price,
            "seat_name": seat_name,
            "seat_multiplier": seat_multiplier,
            "booking_factor": booking_factor,
            "final_price": final_price,
        }
    }

    return render(request, "payment_success.html", context)



def payment_cancel(request, booking_id):
    return render(request, "cancel.html", {"message": "Payment canceled!"})


def payment_confirmation(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    # Prevent duplicate payments
    if booking.status == "Confirmed":
        messages.info(request, f"Booking #{booking.id} is already confirmed and paid.")
        return redirect("booking")

    # Confirm the booking
    booking.status = "Confirmed"
    booking.save()

    # Record the payment
    Payment.objects.create(
        booking=booking,
        amount=getattr(booking, "total_amount", 0),  # fallback if no total_amount
        method="Manual",
        status="Completed"
    )

    messages.success(request, f"Booking #{booking.id} has been confirmed and paid.")
    return redirect("booking")

from decimal import Decimal
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from .models import Booking, BookingDetail, Payment

from decimal import Decimal
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from .models import Booking, BookingDetail, Payment


def create_checkout(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    # Mark booking as confirmed
    booking.status = "Confirmed"
    booking.save()

    # Fetch all booking details
    details = booking.details.all()

    # Calculate total price (loop for one_way, round_trip, multi_city)
    total_price = Decimal("0.00")
    calc_breakdown = []

    for detail in details:
        schedule_date = detail.schedule.departure_time.date()
        base_price = detail.schedule.flight.route.base_price
        seat_name = detail.seat.seat_class.name if detail.seat else "Economy"
        seat_multiplier = detail.seat.seat_class.price_multiplier if detail.seat else Decimal("1.0")
        days_diff = (schedule_date - booking.created_at.date()).days

        # booking factor
        if days_diff >= 30:
            booking_factor = Decimal("0.8")
        elif 7 <= days_diff <= 29:
            booking_factor = Decimal("1.0")
        else:
            booking_factor = Decimal("1.5")

        final_price = base_price * seat_multiplier * booking_factor
        detail.price = final_price
        detail.save()

        total_price += final_price

        calc_breakdown.append({
            "schedule": detail.schedule,
            "seat_name": seat_name,
            "base_price": base_price,
            "seat_multiplier": seat_multiplier,
            "days_diff": days_diff,
            "booking_factor": booking_factor,
            "final_price": final_price,
        })

    # Create payment record
    payment = Payment.objects.create(
        booking=booking,
        amount=total_price,
        method="Manual",
        status="Completed"
    )

    # Send confirmation email
    if booking.student.email:
        subject = f"Booking Confirmed - {booking.id}"
        message = render_to_string("booking_confirmation.html", {
            "booking": booking,
            "payment": payment,
            "details": details,
            "calc_breakdown": calc_breakdown,
        })
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[booking.student.email]
        )
        email.content_subtype = "html"
        email.send(fail_silently=True)

    return render(request, "payment_success.html", {
        "booking": booking,
        "payment": payment,
        "calc_breakdown": calc_breakdown,
        "details": details
    })

